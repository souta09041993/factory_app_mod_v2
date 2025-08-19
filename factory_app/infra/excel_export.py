
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from .logging_conf import get_logger
from ..domain.state import CIRCLE

LOGGER = get_logger(__name__)

def export_monthly_check_sheet(
    *, year: int, month: int, machine: str, type_name: str,
    items: List[str], matrix: Dict[str, List[str]], signs: List[str], save_path: str
) -> str:
    """
    items: 行ラベル
    matrix: {item: ["", "○", ...]}
    signs: 各日のサイン（長さ = 月日数）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{type_name}"[:31]

    # ヘッダ
    ws.cell(1, 1, f"{type_name}（{year}年{month}月） 機械: {machine}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2+len(signs))
    ws.cell(1,1).font = Font(bold=True)
    ws.cell(1,1).alignment = Alignment(horizontal="left")

    # 2行目: 日付ヘッダ
    ws.cell(2, 1, "点検項目")
    for d in range(len(signs)):
        ws.cell(2, 2+d, f"{d+1}")
        ws.cell(2, 2+d).alignment = Alignment(horizontal="center")

    # 本体
    yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    r = 3
    for it in items:
        ws.cell(r, 1, it)
        row = matrix.get(it, [])
        for d, val in enumerate(row, start=0):
            c = ws.cell(r, 2+d, val)
            c.alignment = center
            if val != CIRCLE:
                c.fill = yellow  # 未実施は黄色
        r += 1

    # サイン行
    ws.cell(r, 1, "サイン")
    for d, s in enumerate(signs, start=0):
        c = ws.cell(r, 2+d, s)
        c.alignment = center

    # 幅調整
    ws.column_dimensions["A"].width = 22
    for i in range(2, 2+len(signs)):
        ws.column_dimensions[chr(64+i)].width = 4

    wb.save(save_path)
    LOGGER.info(f"Excel exported -> {save_path}")
    return save_path
